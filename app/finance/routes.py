"""
Finance & Accounting API Routes
"""
from fastapi import APIRouter, HTTPException, status, Depends, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime
import io

from app.utils.auth import get_current_user
from app.finance.models import (
    ChartOfAccountCreate,
    ChartOfAccountUpdate,
    JournalEntryCreate,
    ManualEntryCreate,
)
from app.finance import services, reports
from app.finance.journal_engine import create_journal_entry, validate_double_entry

router = APIRouter(prefix="/finance", tags=["Finance & Accounting"])


# ═══════════════════════════════════════════════════════════════════════════════
# Chart of Accounts
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/coa", status_code=status.HTTP_201_CREATED)
async def create_account(
    payload: ChartOfAccountCreate,
    current_user: dict = Depends(get_current_user),
):
    data = payload.model_dump()
    if not data.get("organization_id"):
        data["organization_id"] = (
            current_user.get("organization_id") or current_user.get("sub")
        )
    try:
        return await services.create_account(data, _user(current_user))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/coa")
async def list_accounts(
    organization_id: Optional[str] = None,
    account_type: Optional[str] = None,
    is_active: Optional[bool] = None,
    current_user: dict = Depends(get_current_user),
):
    org_id = organization_id or current_user.get("organization_id") or current_user.get("sub")
    return await services.get_accounts(org_id, account_type, is_active)


@router.put("/coa/{account_id}")
async def update_account(
    account_id: str,
    payload: ChartOfAccountUpdate,
    current_user: dict = Depends(get_current_user),
):
    data = payload.model_dump(exclude_unset=True)
    try:
        return await services.update_account(account_id, data, _user(current_user))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/coa/seed/{organization_id}", status_code=status.HTTP_201_CREATED)
async def seed_coa(
    organization_id: str,
    current_user: dict = Depends(get_current_user),
):
    """Seed default Chart of Accounts for an organisation (idempotent)."""
    result = await services.seed_chart_of_accounts(organization_id, _user(current_user))
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Journal Entries
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/journal")
async def list_journal_entries(
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    reference_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user),
):
    org_id = organization_id or current_user.get("organization_id")
    role   = current_user.get("role")
    if role == "agency":
        agency_id = agency_id or current_user.get("agency_id") or current_user.get("sub")
    elif role == "branch":
        branch_id = branch_id or current_user.get("branch_id") or current_user.get("sub")

    return await services.get_journal_entries(
        org_id, branch_id, agency_id, reference_type, date_from, date_to, skip, limit
    )


@router.get("/journal/{entry_id}")
async def get_journal_entry(
    entry_id: str,
    current_user: dict = Depends(get_current_user),
):
    entry = await services.get_journal_entry(entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    return entry


@router.post("/journal", status_code=status.HTTP_201_CREATED)
async def post_journal_entry(
    payload: JournalEntryCreate,
    current_user: dict = Depends(get_current_user),
):
    """Post a custom (manually-constructed) journal entry."""
    entries = [e.model_dump() for e in payload.entries]
    if not validate_double_entry(entries):
        raise HTTPException(
            status_code=400,
            detail="Double-entry violation: total debits must equal total credits."
        )
    org_id = payload.organization_id or current_user.get("organization_id")
    try:
        return await create_journal_entry(
            reference_type=payload.reference_type,
            reference_id=payload.reference_id or "",
            description=payload.description,
            entries=entries,
            organization_id=org_id,
            branch_id=payload.branch_id,
            agency_id=payload.agency_id,
            created_by=_user(current_user),
            date_str=payload.date,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/journal/{entry_id}", status_code=status.HTTP_200_OK)
async def reverse_journal_entry(
    entry_id: str,
    current_user: dict = Depends(get_current_user),
):
    try:
        return await services.delete_journal_entry(entry_id, _user(current_user))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# Manual Entry
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/manual-entry", status_code=status.HTTP_201_CREATED)
async def create_manual_entry(
    payload: ManualEntryCreate,
    current_user: dict = Depends(get_current_user),
):
    data = payload.model_dump()
    if not data.get("organization_id"):
        data["organization_id"] = current_user.get("organization_id")
    try:
        return await services.create_manual_entry(data, _user(current_user))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# Reports
# ═══════════════════════════════════════════════════════════════════════════════

def _report_params(current_user: dict) -> dict:
    return {
        "organization_id": current_user.get("organization_id"),
        "branch_id":       current_user.get("branch_id"),
        "agency_id":       (
            current_user.get("agency_id") or
            (current_user.get("sub") if current_user.get("role") == "agency" else None)
        ),
    }


@router.get("/reports/dashboard")
async def finance_dashboard(
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    base = _report_params(current_user)
    org  = organization_id or base["organization_id"]
    br   = branch_id       or base["branch_id"]
    ag   = agency_id       or base["agency_id"]
    return await reports.get_dashboard_kpis(org, br, ag, date_from, date_to)


@router.get("/reports/profit-loss")
async def profit_and_loss(
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    base = _report_params(current_user)
    return await reports.get_profit_and_loss(
        organization_id or base["organization_id"],
        branch_id       or base["branch_id"],
        agency_id       or base["agency_id"],
        date_from, date_to,
    )


@router.get("/reports/balance-sheet")
async def balance_sheet(
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    base = _report_params(current_user)
    return await reports.get_balance_sheet(
        organization_id or base["organization_id"],
        branch_id       or base["branch_id"],
        agency_id       or base["agency_id"],
        date_from, date_to,
    )


@router.get("/reports/trial-balance")
async def trial_balance(
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    base = _report_params(current_user)
    return await reports.get_trial_balance(
        organization_id or base["organization_id"],
        branch_id       or base["branch_id"],
        agency_id       or base["agency_id"],
        date_from, date_to,
    )


@router.get("/reports/ledger")
async def ledger(
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    account_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user),
):
    base = _report_params(current_user)
    return await reports.get_ledger(
        organization_id or base["organization_id"],
        branch_id       or base["branch_id"],
        agency_id       or base["agency_id"],
        account_id, date_from, date_to, skip, limit,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Downloadable Reports (Excel / PDF)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/reports/download/excel/{report_type}")
async def download_excel(
    report_type: str,
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Download any report as Excel. report_type: profit-loss | balance-sheet | trial-balance | ledger"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    base = _report_params(current_user)
    org  = organization_id or base["organization_id"]
    br   = branch_id       or base["branch_id"]
    ag   = agency_id       or base["agency_id"]

    data = await _fetch_report(report_type, org, br, ag, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace("-", " ").title()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    def _write_section(title: str, rows: list, cols: list, start_row: int) -> int:
        # Section title
        ws.cell(row=start_row, column=1).value = title
        ws.cell(row=start_row, column=1).font  = Font(bold=True, size=12)
        start_row += 1
        # Headers
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=start_row, column=ci)
            cell.value = col
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center")
        start_row += 1
        # Rows
        for row in rows:
            for ci, col in enumerate(cols, 1):
                ws.cell(row=start_row, column=ci).value = row.get(col.lower().replace(" ", "_"))
            start_row += 1
        return start_row + 1

    current_row = 1
    ws.cell(row=current_row, column=1).value = f"Report: {ws.title}"
    ws.cell(row=current_row, column=1).font  = Font(bold=True, size=14)
    current_row += 1
    ws.cell(row=current_row, column=1).value = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    current_row += 2

    if report_type == "trial-balance":
        cols = ["Account Code", "Account Name", "Account Type", "Total Debit", "Total Credit", "Balance Debit", "Balance Credit"]
        ws.cell(row=current_row, column=1).value = "Trial Balance"
        ws.cell(row=current_row, column=1).font  = Font(bold=True, size=12)
        current_row += 1
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=current_row, column=ci)
            cell.value = col; cell.font = header_font; cell.fill = header_fill
        current_row += 1
        for r in data.get("rows", []):
            ws.cell(row=current_row, column=1).value = r.get("account_code")
            ws.cell(row=current_row, column=2).value = r.get("account_name")
            ws.cell(row=current_row, column=3).value = r.get("account_type")
            ws.cell(row=current_row, column=4).value = r.get("total_debit")
            ws.cell(row=current_row, column=5).value = r.get("total_credit")
            ws.cell(row=current_row, column=6).value = r.get("balance_debit")
            ws.cell(row=current_row, column=7).value = r.get("balance_credit")
            current_row += 1

    elif report_type == "profit-loss":
        for section, lbl in [("income", "Income"), ("expenses", "Expenses")]:
            ws.cell(row=current_row, column=1).value = lbl
            ws.cell(row=current_row, column=1).font  = Font(bold=True, size=12)
            current_row += 1
            for r in data.get(section, []):
                ws.cell(row=current_row, column=1).value = r.get("account_name")
                ws.cell(row=current_row, column=2).value = r.get("net")
                current_row += 1
            current_row += 1
        ws.cell(row=current_row, column=1).value = "Gross Profit"
        ws.cell(row=current_row, column=2).value = data.get("gross_profit")
        current_row += 1
        ws.cell(row=current_row, column=1).value = "Net Profit"
        ws.cell(row=current_row, column=2).value = data.get("net_profit")

    else:
        # Generic row export
        ws.cell(row=current_row, column=1).value = str(data)

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/reports/download/pdf/{report_type}")
async def download_pdf(
    report_type: str,
    organization_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    agency_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Download any report as PDF. Requires reportlab."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed. Run: pip install reportlab")

    base = _report_params(current_user)
    org  = organization_id or base["organization_id"]
    br   = branch_id       or base["branch_id"]
    ag   = agency_id       or base["agency_id"]

    data = await _fetch_report(report_type, org, br, ag, date_from, date_to)
    stream = io.BytesIO()
    doc    = SimpleDocTemplate(stream, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elems  = []

    title = report_type.replace("-", " ").title()
    elems.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elems.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elems.append(Spacer(1, 0.5 * cm))

    HEADER_STYLE = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
    ])

    if report_type == "trial-balance":
        headers = ["Code", "Account", "Type", "Dr", "Cr", "Bal Dr", "Bal Cr"]
        table_data = [headers] + [
            [r.get("account_code"), r.get("account_name"), r.get("account_type"),
             r.get("total_debit"), r.get("total_credit"), r.get("balance_debit"), r.get("balance_credit")]
            for r in data.get("rows", [])
        ]
        t = Table(table_data, repeatRows=1)
        t.setStyle(HEADER_STYLE)
        elems.append(t)

    elif report_type == "profit-loss":
        for section, lbl in [("income", "Income"), ("expenses", "Expenses")]:
            elems.append(Paragraph(f"<b>{lbl}</b>", styles["Heading2"]))
            rows_data = [["Account", "Amount"]] + [
                [r.get("account_name"), r.get("net")] for r in data.get(section, [])
            ]
            t = Table(rows_data, repeatRows=1)
            t.setStyle(HEADER_STYLE)
            elems.append(t)
            elems.append(Spacer(1, 0.3 * cm))
        summary = [["", ""], ["Gross Profit", data.get("gross_profit")], ["Net Profit", data.get("net_profit")]]
        t = Table(summary)
        t.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold")]))
        elems.append(t)

    else:
        elems.append(Paragraph(str(data), styles["Normal"]))

    doc.build(elems)
    stream.seek(0)
    filename = f"{report_type}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        stream,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


async def _fetch_report(report_type, org, br, ag, date_from, date_to) -> dict:
    if report_type == "profit-loss":
        return await reports.get_profit_and_loss(org, br, ag, date_from, date_to)
    elif report_type == "balance-sheet":
        return await reports.get_balance_sheet(org, br, ag, date_from, date_to)
    elif report_type == "trial-balance":
        return await reports.get_trial_balance(org, br, ag, date_from, date_to)
    elif report_type == "ledger":
        return await reports.get_ledger(org, br, ag, None, date_from, date_to)
    else:
        raise HTTPException(status_code=400, detail=f"Unknown report type: {report_type}")


# ═══════════════════════════════════════════════════════════════════════════════
# Audit Trail
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/audit-trail")
async def audit_trail(
    action: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user: dict = Depends(get_current_user),
):
    return await services.get_audit_trail(
        organization_id=current_user.get("organization_id"),
        action=action,
        skip=skip,
        limit=limit,
    )


# ─── util ──────────────────────────────────────────────────────────────────────

def _user(current_user: dict) -> str:
    return (
        current_user.get("email") or
        current_user.get("username") or
        current_user.get("sub", "system")
    )
